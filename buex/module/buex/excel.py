import requests
import json
from datetime import datetime, timedelta
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import os.path
import random
import time
from buex.module.buex import common as mcm

def render_excel(context, excel_filename):
    if context['contentAlias'] == 'monthlyreport':
        context['excelcontent'] = exportmonthlyreportcontent(context)
        render_monthlyreport_excel(context, excel_filename)

def get_excel_defaul_key(index):
    if index < 0:
        return 'A'
    firstletter_index = index // 26
    secondletter_index = index % 26
    if index >= 0 and index <= 701:
        unicode = secondletter_index + 65
        secondletter = chr(unicode)
        if firstletter_index == 0:
            firstletter = ''
        elif firstletter_index > 0:
            unicode = (firstletter_index - 1) + 65
            firstletter = chr(unicode)
        return firstletter + secondletter
    elif index > 701:
        return None

def set_width_adjustment(ws, type='', letter={}):
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter  # Get the column name
        for cell in col:
            if cell.coordinate in ws.merged_cells:  # not check merge_cells
                continue
            try:  # Necessary to avoid error on empty cells
                # max length calc

                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
                # style setting
                if type == 'shippingadjust':
                    # cell.alignment = Alignment(horizontal='center')
                    pass
            except:
                pass
        if type == 'shippingadjust':
            adjusted_width = max_length + 7
        elif type == 'exportorder':
            adjusted_width = max_length + 7
            if adjusted_width >= 30:
                adjusted_width = 30
            if letter['th_letter'] == column:
                adjusted_width = 50
        else:
            adjusted_width = max_length + 7
            if adjusted_width >= 30:
                adjusted_width = 30
        ws.column_dimensions[column].width = adjusted_width

def exportmonthlyreportcontent(context):
    try:
        request = context['request']
        params = context['param']
        monthlyreportlist = get_monthlyreportlist(request, params)
    except:
        monthlyreportlist = []

    headerinfo = [
        {'field': 'tracking_number', 'fieldalias': 'Tracking Number'},
        {'field': 'ship_date', 'fieldalias': 'Shipped At'},
    ]
    return {'monthlyreportlist': monthlyreportlist, 'headers': headerinfo}

def get_monthlyreportlist(request, params):
    current_year = params['current_year']
    monthlyreportlist = mcm.get_monthly_report(current_year)
    return monthlyreportlist

def render_monthlyreport_excel(context, excel_filename):
    wb = openpyxl.Workbook()
    wb.save(excel_filename)
    csv_file = openpyxl.load_workbook(excel_filename)
    sheet = csv_file['Sheet']
    sheet.freeze_panes = "A2"
    index = 1
    for k in range(0, len(context['excelcontent']['headers'])):
        letter = get_excel_defaul_key(k)
        sheet[letter + str(index)] = context['excelcontent']['headers'][k]['fieldalias']
        sheet[letter + str(index)].fill = PatternFill(fgColor="8C8C8C", fill_type="solid")
        sheet[letter + str(index)].alignment = Alignment(horizontal='center')
    for i in range(0, len(context['excelcontent']['exportuserinfolist'])):
        index += 1
        for j in range(0, len(context['excelcontent']['headers'])):
            letter = get_excel_defaul_key(j)
            sheet[letter + str(index)] = context['excelcontent']['exportuserinfolist'][i][
                context['excelcontent']['headers'][j]['field']]
            sheet[letter + str(index)].alignment = Alignment(horizontal='center')

    FullRange = "A1:" + get_column_letter(sheet.max_column) + str(sheet.max_row)
    sheet.auto_filter.ref = FullRange
    set_width_adjustment(sheet, type='')
    csv_file.save(excel_filename)

